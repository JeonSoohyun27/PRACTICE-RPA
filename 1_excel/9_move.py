from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

#번호 영어 수학
#번호 국어 영어 수학 의 형태로 이동
ws.move_range("B1:C11",rows=0,cols=1) #영어 수학의 범위 행은 움직이지않고 열은 오른쪽으로 1칸
ws['B1'].value = "국어" #비워진 B1셀에 '국어' 입력

#번호 영어 수학
#번호 수학 영어 의 형태로 이동
ws.move_range("C1:C11",rows=0,cols=-1) #-1로 좌로 움직여줌

wb.save("sample_move.xlsx")