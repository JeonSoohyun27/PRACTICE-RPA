from openpyxl import load_workbook # workbook을 불러와줄때

wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row = 2):
    # 번호, 영어, 수학
    if int(row[1].value)>70:
        print(row[0].value,"번 학생은 영어 고득점자")
        

#"영어"가 아닌 "컴퓨터"로 내용을 바꿔줄때 
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")