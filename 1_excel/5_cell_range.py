from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호","영어","수학"]) # A, B, C
for i in range(1,11): # 10개 데이터 넣기
    ws.append([i,randint(0,100),randint(0,100)])

# col_B = ws["B"] # 하나의(영어) column만 가지고 오기
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"] # 'B부터 C까지' 여러(영어,수학) column 함께 가지고오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1] # 첫번째 row만 가지고 오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# row_range1 = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range1:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print('---')


# #좌표 cell 주소가 필요할 때 
# from openpyxl.utils.cell import coordinate_from_string

# row_range2 = ws[2:ws.max_row] #2번째 줄부터 마지막 줄까지
# for rows in row_range2:
#     for cell in rows:
#         #print(cell.coordinate, end=" ") # A/10, AZ/250
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy[0],end="") # A
#         print(xy[1],end=" ") # 1
#     print('###')


# # 전체 rows 가져오기
# for row in ws.iter_rows():
#     print(row[2].value)

# # 전체 column 가져오기
# for column in ws.iter_cols():
#     print(column[2].value)


# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2,max_row=11,min_col=2,max_col=3):
    print(row[0].value,row[1].value) #수학,영어


wb.save("sample.xlsx")